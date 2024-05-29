from openpyxl import load_workbook

# 读取xlsx文件
def readxlsxfile(path):
    workbook = load_workbook(path)
    username=[]
    password=[]
    # 获取所有表的名称
    sheet_names = workbook.sheetnames

    # 遍历每个表
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        
        # 获取表的行数和列数
        max_row = sheet.max_row
        max_column = sheet.max_column
        
        # 遍历每行
        for row in range(1, max_row + 1):
            row_data = []
            username.append(str(sheet.cell(row=row, column=1).value))
            # 遍历每列
            password.append(str(sheet.cell(row=row, column=2).value))
    # 关闭工作簿
    workbook.close()
    return username,password
def readquizfile(path):
    workbook = load_workbook(path)
    # 获取所有表的名称
    {"type": "choice", "question": "1. 下列选项中不属于函数优点的是（  ）。", "choices": ["A. 减少代码重复", "B. 使程序模块化", "C. 使程序便于阅读", "D. 便于发挥程序员的创造力"], "answer": "D","explanation":""},
    # 遍历每个表
    ceyansheet=workbook["测验信息"]
    choicesheet=workbook["选择题"]
    tiankongsheet=workbook["填空题"]
    judgesheet=workbook["判断题"]
    name=str(ceyansheet.cell(row=1, column=2).value)
    time=int(ceyansheet.cell(row=2, column=2).value)
    question=[]
    
    max_row = choicesheet.max_row
    for i in range(3,max_row+1):
        data={"type": "","number":0, "question": "", "choices": [], "answer": "D","explanation":"","grade":5}
        data['type']="choice"
        data["number"]=choicesheet.cell(row=i,column=1).value
        data["question"]=choicesheet.cell(row=i,column=2).value
        choice=[]
        choice.append("A. "+choicesheet.cell(row=i,column=3).value)
        choice.append("B. "+choicesheet.cell(row=i,column=4).value)
        choice.append("C. "+choicesheet.cell(row=i,column=5).value)
        choice.append("D. "+choicesheet.cell(row=i,column=6).value)
        data["choices"]=choice
        data["answer"]=str(choicesheet.cell(row=i,column=7).value)
        data["explanation"]=choicesheet.cell(row=i,column=8).value
        question.append(data)
    
    max_row = tiankongsheet.max_row

    for i in range(3,max_row+1):
        data={"type": "","number":0, "question": "", "choices": [], "answer": "D","explanation":"","grade":5}
        data["type"]="fill_in"
        data["number"]=tiankongsheet.cell(row=i,column=1).value
        data["question"]=tiankongsheet.cell(row=i,column=2).value
        data["answer"]=tiankongsheet.cell(row=i,column=3).value
        data["explanation"]=tiankongsheet.cell(row=i,column=4).value
        question.append(data)
    workbook.close()
    return name,time,question
